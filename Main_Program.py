import subprocess
import sys
import os
import re
from glob import glob
import pandas as pd
from tkinter import END, Tk, StringVar, filedialog, messagebox, DISABLED, NORMAL
from tkinter import ttk, scrolledtext
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import threading

# Function to check and install required libraries
def install_requirements():
    required_libraries = ['pandas', 'openpyxl']
    for library in required_libraries:
        try:
            __import__(library)
        except ImportError:
            print(f"{library} not found, installing...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", library])

# Ensure all required libraries are installed before running the rest of the script
install_requirements()

def extract_range_from_filename(filename):
    """
    Extract numeric ranges from the filename for sorting.
    Expected filename format: something_<start>_<end>.xlsx
    """
    match = re.search(r'_(\d+)_(\d+)\.xlsx$', filename)
    if match:
        return (int(match.group(1)), int(match.group(2)))
    return (0, 0)

def combine_files(folder_path, output_file, log_function, progress_callback):
    """
    Combine Excel files from the specified folder into a single workbook.
    """
    try:
        # Initialize workbook and sheet names
        log_function("Initializing combined workbook and sheets...")
        combined_wb = Workbook()
        combined_sheets = {
            'Company Details': combined_wb.active,
            'Financial Info': combined_wb.create_sheet(title='Financial Info'),
            'Executive': combined_wb.create_sheet(title='Executive')
        }
        combined_sheets['Company Details'].title = 'Company Details'
       
        # Flag to track if headers have been written
        headers_written = {sheet_name: False for sheet_name in combined_sheets}
       
        # Get a list of all Excel files in the folder and sort them
        excel_files = glob(os.path.join(folder_path, '*.xlsx'))
        excel_files_sorted = sorted(excel_files, key=lambda x: extract_range_from_filename(os.path.basename(x)))
        log_function(f"Found {len(excel_files_sorted)} Excel files in the folder: {folder_path}")
       
        if not excel_files_sorted:
            log_function("No Excel files found in the selected folder.")
            return False
       
        total_files = len(excel_files_sorted)
        for index, file in enumerate(excel_files_sorted, start=1):
            log_function(f"Processing file: {os.path.basename(file)}")
            workbook = pd.ExcelFile(file)
           
            # Sheets to process
            sheets_to_process = ['Company Details', 'Financial Info', 'Executive']
           
            for sheet_name in sheets_to_process:
                if sheet_name in workbook.sheet_names:
                    log_function(f"Copying worksheet: {sheet_name}")
                    sheet_data = pd.read_excel(workbook, sheet_name)
                    sheet = combined_sheets[sheet_name]
                    rows = dataframe_to_rows(sheet_data, index=False, header=not headers_written[sheet_name])
                   
                    for row in rows:
                        sheet.append(row)
                   
                    headers_written[sheet_name] = True
                    log_function(f"Finished copying worksheet: {sheet_name}")
       
            # Update progress after each file
            progress_callback(index)
       
        # Save the combined workbook
        combined_wb.save(output_file)
        log_function(f"Combined workbook saved as: {output_file}")
        return True
    except Exception as e:
        log_function(f"Error: {str(e)}")
        return False

class ExcelCombinerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel File Combiner")
        self.root.geometry("800x700")
        self.root.resizable(False, False)
       
        # Set dark theme styles
        self.style = ttk.Style()
        self.style.theme_use('clam')  # Using 'clam' as base for customization
        self.style.configure('.', background='#2e2e2e', foreground='#ffffff',
                             font=('Segoe UI', 10))
        self.style.configure('TButton', background='#444444', foreground='#ffffff',
                             relief='flat')
        self.style.configure('TLabel', background='#2e2e2e', foreground='#ffffff')
        self.style.configure('TEntry', fieldbackground='#3e3e3e', foreground='#ffffff')
        self.style.map('TButton', background=[('active', '#555555')])
        self.style.configure('TFrame', background='#2e2e2e')
        self.style.configure("Horizontal.TProgressbar", troughcolor="#3e3e3e",
                             background="#4caf50")
       
        # Variables
        self.input_folder = StringVar()
        self.output_file = StringVar()
       
        # UI Setup
        self.setup_ui()
   
    def setup_ui(self):
        # Frame for input folder selection
        input_frame = ttk.Frame(self.root, padding=10)
        input_frame.pack(fill='x')
       
        input_label = ttk.Label(input_frame, text="Input Folder:")
        input_label.pack(side='left', padx=(0, 10))
       
        self.input_entry = ttk.Entry(input_frame, textvariable=self.input_folder, width=50, state='readonly')
        self.input_entry.pack(side='left', fill='x', expand=True)
       
        browse_input_btn = ttk.Button(input_frame, text="Browse", command=self.browse_input_folder)
        browse_input_btn.pack(side='left', padx=(10, 0))
       
        # Frame for output file selection
        output_frame = ttk.Frame(self.root, padding=10)
        output_frame.pack(fill='x')
       
        output_label = ttk.Label(output_frame, text="Output File:")
        output_label.pack(side='left', padx=(0, 14))
       
        self.output_entry = ttk.Entry(output_frame, textvariable=self.output_file, width=50, state='readonly')
        self.output_entry.pack(side='left', fill='x', expand=True)
       
        browse_output_btn = ttk.Button(output_frame, text="Browse", command=self.browse_output_file)
        browse_output_btn.pack(side='left', padx=(10, 0))
       
        # Frame for list of loaded files
        files_frame = ttk.Frame(self.root, padding=10)
        files_frame.pack(fill='both', expand=True)
       
        files_label = ttk.Label(files_frame, text="Loaded Files:")
        files_label.pack(anchor='w')
       
        self.files_listbox = scrolledtext.ScrolledText(files_frame, height=15, bg='#3e3e3e', fg='#ffffff',
                                                      insertbackground='white', state='disabled', wrap='word')
        self.files_listbox.pack(fill='both', expand=True, pady=(5, 10))
       
        # Combine button
        combine_btn = ttk.Button(self.root, text="Combine Files", command=self.start_combining)
        combine_btn.pack(pady=(10, 5))
       
        # Progress Bar
        self.progress = ttk.Progressbar(self.root, orient='horizontal', mode='determinate',
                                        length=700, style="Horizontal.TProgressbar")
        self.progress.pack(pady=(0, 10))
        
        # Spinner (Indeterminate Progressbar)
        self.spinner = ttk.Progressbar(self.root, mode='indeterminate')
        self.spinner.pack(pady=(0, 10))
        self.spinner.pack_forget()  # Hide initially
        
        # Status bar
        self.status_var = StringVar()
        self.status_var.set("Select input and output locations to begin.")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief='sunken', anchor='w')
        status_bar.pack(fill='x', side='bottom')
   
    def browse_input_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.input_folder.set(folder_selected)
            self.log_files(folder_selected)
            self.status_var.set(f"Selected input folder: {folder_selected}")
           
            # Ask the user if they want to clean and convert file names
            clean = messagebox.askyesno("Clean and Convert File Names",
                                        "Do you want to convert and clean up file names in the selected folder?")
            if clean:
                # Run the file cleaning process in a separate thread to keep UI responsive
                threading.Thread(target=self.clean_file_names, args=(folder_selected,)).start()
   
    def browse_output_file(self):
        file_selected = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                    filetypes=[("Excel files", "*.xlsx")],
                                                    title="Select Output File")
        if file_selected:
            self.output_file.set(file_selected)
            self.status_var.set(f"Selected output file: {file_selected}")
   
    def log_files(self, folder_path):
        """
        Display the list of Excel files in the selected input folder in sorted order.
        """
        self.files_listbox.config(state=NORMAL)
        self.files_listbox.delete(1.0, END)
        excel_files = glob(os.path.join(folder_path, '*.xlsx'))
        if not excel_files:
            self.files_listbox.insert(END, "No Excel files found in the selected folder.\n")
        else:
            # Sort the files using the extract_range_from_filename function
            excel_files_sorted = sorted(excel_files, key=lambda x: extract_range_from_filename(os.path.basename(x)))
            for file in excel_files_sorted:
                self.files_listbox.insert(END, os.path.basename(file) + "\n")
        self.files_listbox.config(state=DISABLED)
   
    def log_message(self, message):
        """
        Log messages to the files_listbox and update status.
        """
        self.files_listbox.config(state=NORMAL)
        self.files_listbox.insert(END, message + "\n")
        self.files_listbox.see('end')
        self.files_listbox.config(state=DISABLED)
        self.status_var.set(message)
   
    def update_progress(self, current):
        """
        Update the progress bar with the current progress.
        """
        self.progress['value'] = current
        self.root.update_idletasks()  # Refresh the UI to reflect the change
   
    def start_combining(self):
        """
        Start the combining process in a separate thread to keep the UI responsive.
        """
        self.spinner.pack(pady=(0, 10))  # Show spinner
        self.spinner.start(10)  # Start spinner with a delay of 10ms
        thread = threading.Thread(target=self.combine)
        thread.start()
   
    def combine(self):
        try:
            input_dir = self.input_folder.get()
            output_path = self.output_file.get()
        
            if not input_dir:
                messagebox.showerror("Error", "Please select an input folder.")
                return
            if not output_path:
                messagebox.showerror("Error", "Please select an output file.")
                return
        
            # Get the list of Excel files to determine progress
            excel_files = glob(os.path.join(input_dir, '*.xlsx'))
            excel_files_sorted = sorted(excel_files, key=lambda x: extract_range_from_filename(os.path.basename(x)))
            total_files = len(excel_files_sorted)
        
            if total_files == 0:
                messagebox.showerror("Error", "No Excel files found in the selected folder.")
                return
        
            # Configure the progress bar
            self.progress['maximum'] = total_files
            self.progress['value'] = 0
        
            # Confirm overwrite if output file exists
            if os.path.exists(output_path):
                overwrite = messagebox.askyesno("Overwrite Confirmation",
                                            f"The file '{output_path}' already exists. Do you want to overwrite it?")
                if not overwrite:
                    self.progress['value'] = 0
                    return
        
            # Disable UI elements during processing
            self.disable_ui()
            self.files_listbox.config(state=NORMAL)
            self.files_listbox.delete(1.0, END)
            self.files_listbox.config(state=DISABLED)
        
            self.log_message("Starting the combining process...")
        
            # Run the combine_files function
            success = combine_files(input_dir, output_path, self.log_message, self.update_progress)
        
 # After successful combination
            if success:
                messagebox.showinfo("Success", f"Files have been successfully combined into:\n{output_path}")
                self.status_var.set("Combination successful.")
                self.log_message("Combination process completed successfully.")
                self.log_files(input_dir)  # Refresh the files list
            else:
                messagebox.showerror("Failed", "An error occurred during the combining process. Check logs for details.")
                self.status_var.set("Combination failed.")
        finally:
            # Stop and hide the spinner
            self.spinner.stop()
            self.spinner.pack_forget()
            # Reset the progress bar
            self.progress['value'] = 0
            # Re-enable UI elements
            self.enable_ui()
   
   
    def disable_ui(self):
        """
        Disable UI elements to prevent interaction during processing.
        Only disables widgets that support the 'state' option.
        """
        for child in self.root.winfo_children():
            if isinstance(child, ttk.Frame):
                for subchild in child.winfo_children():
                    if isinstance(subchild, (ttk.Button, ttk.Entry, scrolledtext.ScrolledText)):
                        subchild.configure(state='disabled')
            elif isinstance(child, ttk.Button):
                child.configure(state='disabled')
        # Additionally disable the status bar to prevent changes
        # Note: Labels generally don't have a 'state' option, hence no action needed
   
    def enable_ui(self):
        """
        Enable UI elements after processing.
        Only enables widgets that support the 'state' option.
        """
        for child in self.root.winfo_children():
            if isinstance(child, ttk.Frame):
                for subchild in child.winfo_children():
                    if isinstance(subchild, ttk.Button):
                        subchild.configure(state='normal')
                    elif isinstance(subchild, ttk.Entry):
                        subchild.configure(state='readonly')
                    elif isinstance(subchild, scrolledtext.ScrolledText):
                        subchild.configure(state='disabled')
            elif isinstance(child, ttk.Button):
                child.configure(state='normal')
   
    def clean_file_names(self, folder_path):
        """
        Clean and convert .csv files to .xlsx in the specified folder.
        Renames files based on specific rules.
        """
        try:
            self.log_message("Starting file name cleaning and conversion...")
            counter_num_files = 0
            # Iterate through all files in the specified folder
            for filename in os.listdir(folder_path):
                if filename.endswith('.csv'):
                    # Construct full file path
                    index = filename.find('_advancesearch')
                    if index != -1:
                        # Create the new filename by slicing the original filename
                        new_filename = filename[:index] + '.xlsx'
                    else:
                        new_filename = filename[:-4] + '.xlsx'  # Remove .csv and add .xlsx
                    csv_file = os.path.join(folder_path, filename)
                   
                    # Create the new file name with .xlsx extension
                    xlsx_file = os.path.join(folder_path, new_filename)
                   
                    # Rename the file
                    os.rename(csv_file, xlsx_file)
                    self.log_message(f"Renamed: {filename} to {new_filename}")
                    counter_num_files += 1
            self.log_message(f"\nThe number of files renamed is: {counter_num_files}\n")
            messagebox.showinfo("File Renaming Completed", f"Renamed {counter_num_files} files successfully.")
            # Refresh the list of files displayed in sorted order
            self.log_files(folder_path)
        except Exception as e:
            self.log_message(f"Error during file renaming: {str(e)}")
            messagebox.showerror("Error", f"An error occurred during file renaming:\n{str(e)}")

def main():
    root = Tk()
    root.configure(bg='#2e2e2e')
    app = ExcelCombinerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
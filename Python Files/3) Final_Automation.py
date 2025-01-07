import pandas as pd
import os
from glob import glob
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import re

# this function sorts the file name in numerical order
def extract_range_from_filename(filename):
    # Use regex to find the numeric range in the filename
    match = re.search(r'_(\d+)_(\d+)\.xlsx$', filename)
    if match:
        return (int(match.group(1)), int(match.group(2)))  # Return tuple of (start, end)
    return (0, 0)  # Default if no match found

def combine_workbooks_with_formatting(folder_path, output_file):
    # Initialize workbook and sheet names
    print("Initializing combined workbook and sheets...")
    combined_wb = Workbook()
    combined_sheets = {
        'Company Details': combined_wb.active,  # Use the default active sheet for the first
        'Financial Info': combined_wb.create_sheet(title='Financial Info'),
        'Executive': combined_wb.create_sheet(title='Executive')
    }
    combined_sheets['Company Details'].title = 'Company Details'  # Rename active sheet
    
    # Flag to track if headers have been written
    headers_written = {sheet_name: False for sheet_name in combined_sheets}

    # Get a list of all Excel files in the folder and sort them based on numeric ranges in their filenames
    excel_files = glob(os.path.join(folder_path, '*.xlsx'))
    excel_files.sort(key=extract_range_from_filename)  # Sort files by numeric range
    print(f"Found {len(excel_files)} Excel files in the folder: {folder_path}")

    # Go through each file
    for file in excel_files:
        print(f"Processing file: {file}")
        # Load each workbook
        workbook = pd.ExcelFile(file)
        
        # Sheets to process
        sheets_to_process = ['Company Details', 'Financial Info', 'Executive']
        
        for sheet_name in sheets_to_process:
            # Check if the sheet exists in the current workbook
            if sheet_name in workbook.sheet_names:
                print(f"Copying worksheet: {sheet_name} from {file}")
                # Read the sheet into a DataFrame
                sheet_data = pd.read_excel(workbook, sheet_name)
                
                # Get the target combined sheet
                sheet = combined_sheets[sheet_name]
                
                # Convert DataFrame to rows for openpyxl compatibility
                rows = dataframe_to_rows(sheet_data, index=False, header=not headers_written[sheet_name])
                
                # Append rows to the combined sheet
                for row in rows:
                    sheet.append(row)
                
                # Mark headers as written for this sheet
                headers_written[sheet_name] = True
                print(f"Finished copying worksheet: {sheet_name}")
    
    # Save the combined workbook
    combined_wb.save(output_file)
    print(f"Combined workbook saved as: {output_file}")

# Use the function
folder_path = 'Retail'  # Replace with the path to your folder with Excel files
output_file = 'Combined_Workbook_With_Formatting.xlsx'  # Desired output file name
combine_workbooks_with_formatting(folder_path, output_file)
